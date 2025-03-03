from datetime import datetime,timedelta
from openpyxl import load_workbook
import pyodbc
import os

now=datetime.now()
now = datetime.now()
inizio_mese = now.replace(day=1).strftime('%Y%m%d')
ieri = (now - timedelta(days=1)).strftime('%Y%m%d')


def get_db_connection():
    return pyodbc.connect(
        driver='{iSeries Access ODBC Driver}',
        system='192.168.100.2',
        uid='BALADMIN9',
        pwd='ADMINBAL'
    )

def execute_query(conn, query):
    with conn.cursor() as cursor:
        cursor.execute(query)
        return cursor.fetchall(), [column[0] for column in cursor.description]

def create_excel_sheet(wb, sheet_name, data, columns):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    for row_num, row in enumerate(data, 2):
        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=str(cell_value))
    
    if ws['A2'].value:
        ws.sheet_properties.tabColor = 'FF0000'

def analisi_pf():
    cartella='C:/Users/fabrizio_beggiato/Desktop 2/File lavoro Magazzino/analisi inventario pf/'
    file_name = cartella + 'Analisi_PF.xlsx'
    wb = load_workbook(file_name)
    
    with get_db_connection() as conn:
        # Query 1
        query1 = """sELECT 
    cdarmo,
    dsarmo,
    rifemm,
    COALESCE(POSITIVI, 0) AS POSITIVI,
    COALESCE(NEGATIVI, 0) AS NEGATIVI,
    COALESCE(POSITIVI, 0) - COALESCE(NEGATIVI, 0) AS DIFFERENZA
    from(
SELECT
    cdarmo,
    dsarmo,
    rifemm,
    SUM(CASE WHEN camomm IN ('30','31','70','28') THEN qtmomm END) AS POSITIVI,
    SUM(CASE WHEN camomm IN ('35','36','71','57') THEN qtmomm END) AS NEGATIVI
FROM (
    SELECT DISTINCT
        a.cdarmo,
        a.dsarmo,
        a.giatmo,
        qtmomm,
        i.rifemm,
        i.camomm
    FROM
        BAL90dat.mgorc00w AS a
    INNER JOIN
        BAL90dat.mgmov00f AS i ON a.cdarmo = i.cdarmm
    WHERE
        a.giatmo < 0
        AND i.camomm <> '24'
) AS t
GROUP BY
    cdarmo,
    dsarmo,
    rifemm) as aggregated_data
where
	COALESCE(POSITIVI, 0) - COALESCE(NEGATIVI, 0)<>0
order by cdarmo"""  # Inserisci la query completa qui
        data1, columns1 = execute_query(conn, query1)
        create_excel_sheet(wb, 'Db', data1, columns1)
        
        # Query 2
        query2 = "select DTMNS1,nrsps1,tdocs1,nrors1,nrrcs1,cdars1,qte1s1,qte2s1,saacs1 from BAL90DAT.sprig02u where not exists(select * from BAL90DAT.akril00f where tdocs1=tdocrp and nrors1=nrorrp and nrrgs1=nrrgrp and sequrp='860') and exists (select * from BAL90DAT.ocmov01f where tdocs1=tdocoo and nrors1=nroroo and nrrgs1=nrrgoo and timooo='54') and exists (select * from BAL90DAT.spbks00f where nrsps1=nrspsp and dtpcsp<=" + ieri + ") and dt01s1 between " + inizio_mese + " and " + ieri +" order by dtmns1,nrors1, nrrcs1"  # Inserisci la query completa qui
        data2, columns2 = execute_query(conn, query2)
        create_excel_sheet(wb, 'PF_Non_Sparati', data2, columns2)
        
        # Query 3
        query3 = "select * from BAL90dat.pcimp00f where cdfaci='851' and dt01ci>='"+inizio_mese+"'"  # Inserisci la query completa qui
        data3, columns3 = execute_query(conn, query3)
        create_excel_sheet(wb, 'Fase851', data3, columns3)
        
        # Query 4
        query4 = "select pf.DT01PO,pf.ORPRPO,pf.cdarpo,pf.lottpo from BAL90dat.PMORD00F pf  where (NOT EXISTS  (SELECT * FROM BAL90dat.PMIMP00F pf2 WHERE pf.orprpo=pf2.orprpi) OR NOT EXISTS (select * from BAL90dat.PCIMP00F pf3 WHERE pf.orprpo=pf3.orprci))  AND pf.lottpo<>'' AND pf.DT01PO>"+inizio_mese  # Inserisci la query completa qui
        data4, columns4 = execute_query(conn, query4)
        create_excel_sheet(wb, 'ODP_vuoti', data4, columns4)
        
        # Query 5
        query5 = "SELECT concat(nropmm,rigamm)as rigamm,camomm,dtMOMM,cdarmm,B.DSARMA,B.TPSTMA,qtmomm,rifemm,cdmgmm,CMMMMM FROM BAL90DAT.MGMOV00F AS A INNER JOIN BAL90dat.MGART00F AS B ON CDARMM=CDARMA where TPSTMA='1' AND dt01ma>'"+str(now.year)+"0101' AND (CDARMM like '7%' OR CDARMM like 'I%') AND (CDARMM >= '6010000000000' AND CDARMM<= '6999999999999' and cla1ma='63') and camomm='10'"  # Inserisci la query completa qui
        data5, columns5 = execute_query(conn, query5)
        create_excel_sheet(wb, 'Acquisti_Errati', data5, columns5)
    
        # Query 6
        query6 = "SELECT cdarmd FROM BAL90dat.MGART02F mf WHERE CDARMd like '95%' AND cl02md='' AND dtmnmd>='"+inizio_mese+"'"  # Inserisci la query completa qui
        data6, columns6 = execute_query(conn, query6)
        create_excel_sheet(wb, 'Caratt_Modello_Mancante', data6, columns6)
        
    wb.save(file_name)
    print('Aggiornamento eseguito')
    
    os.chdir(cartella)
    os.system('start excel.exe Analisi_PF.xlsx')

    

if __name__ == "__main__":
    analisi_pf()